# 主要作用：运行 Streamlit 页面，维护角色会话、参考文件、建议和版本数据。

from pathlib import Path
from datetime import datetime
import base64
import json
import mimetypes
import ssl
import urllib.error
import urllib.request

import certifi
import openpyxl
import streamlit as st
import streamlit.components.v1 as components

from agents.interface_agent import handle_input


ROOT_DIR = Path(__file__).resolve().parent
FRONTEND_DIR = ROOT_DIR / "frontend"
REFERENCE_DIR = ROOT_DIR / "reference_files"
MANAGEMENT_PATH = ROOT_DIR / "Database/演示互动管理.xlsx"

WEATHER_URL = (
    "https://api.open-meteo.com/v1/forecast"
    "?latitude=22.5431&longitude=114.0579"
    "&current=temperature_2m,weather_code,is_day"
    "&timezone=Asia%2FShanghai"
)

VALID_ROLES = {"A", "B", "C"}


@st.cache_data(ttl=600, show_spinner=False)
def load_weather() -> dict[str, str | int]:
    """读取深圳当前天气；外网不可用时返回静态值。"""
    fallback = {
        "city": "深圳",
        "temperature": 26,
        "weather_code": 1,
        "is_day": 1,
    }

    try:
        context = ssl.create_default_context(
            cafile=certifi.where()
        )

        with urllib.request.urlopen(
            WEATHER_URL,
            context=context,
            timeout=4,
        ) as response:
            current = json.loads(
                response.read().decode("utf-8")
            ).get("current", {})

        return {
            "city": "深圳",
            "temperature": round(
                float(
                    current.get(
                        "temperature_2m",
                        fallback["temperature"],
                    )
                )
            ),
            "weather_code": int(
                current.get(
                    "weather_code",
                    fallback["weather_code"],
                )
            ),
            "is_day": int(
                current.get(
                    "is_day",
                    fallback["is_day"],
                )
            ),
        }

    except (
        urllib.error.URLError,
        TimeoutError,
        OSError,
        ValueError,
        TypeError,
    ):
        return fallback


def load_reference_files() -> list[dict[str, str | int]]:
    """读取参考文件目录，前端不提供文件上传入口。"""
    if not REFERENCE_DIR.exists():
        return []

    allowed_suffixes = {
        ".pdf",
        ".doc",
        ".docx",
        ".xls",
        ".xlsx",
        ".csv",
        ".json",
        ".txt",
        ".md",
        ".ppt",
        ".pptx",
    }

    files = []

    for path in sorted(REFERENCE_DIR.rglob("*")):
        if not path.is_file():
            continue

        if path.name.startswith("."):
            continue

        if path.suffix.lower() not in allowed_suffixes:
            continue

        files.append(
            {
                "name": path.name,
                "path": path.relative_to(
                    REFERENCE_DIR
                ).as_posix(),
                "extension": (
                    path.suffix.lstrip(".").upper()
                    or "FILE"
                ),
                "size": path.stat().st_size,
                "category": (
                    path.parent.relative_to(
                        REFERENCE_DIR
                    ).as_posix()
                    if path.parent != REFERENCE_DIR
                    else "参考文件"
                ),
                "mime_type": (
                    mimetypes.guess_type(path.name)[0]
                    or "application/octet-stream"
                ),
                "data": base64.b64encode(
                    path.read_bytes()
                ).decode("ascii"),
            }
        )

    return files


def load_versions():
    workbook = openpyxl.load_workbook(
        MANAGEMENT_PATH,
        data_only=True,
    )
    sheet = workbook["版本说明"]
    return [
        {
            "version": row[0],
            "description": row[1],
        }
        for row in sheet.iter_rows(
            min_row=2,
            values_only=True,
        )
        if row[0]
    ]


def load_feedback():
    workbook = openpyxl.load_workbook(
        MANAGEMENT_PATH,
        data_only=True,
    )
    sheet = workbook["用户建议"]
    return [
        {
            "feedback_id": row[0],
            "role": row[1],
            "content": row[2],
            "submitted_at": row[3],
            "status": row[4],
            "reply": row[5],
            "replied_at": row[6],
        }
        for row in sheet.iter_rows(
            min_row=2,
            values_only=True,
        )
        if row[0]
    ]


def append_feedback(role, content):
    workbook = openpyxl.load_workbook(
        MANAGEMENT_PATH
    )
    sheet = workbook["用户建议"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    feedback_id = datetime.now().strftime("FB%Y%m%d%H%M%S%f")
    sheet.append(
        [
            feedback_id,
            role,
            content,
            now,
            "待回复",
            "",
            "",
        ]
    )
    workbook.save(MANAGEMENT_PATH)


def reply_feedback(feedback_id, reply):
    workbook = openpyxl.load_workbook(
        MANAGEMENT_PATH
    )
    sheet = workbook["用户建议"]

    for row in sheet.iter_rows(min_row=2):
        if row[0].value == feedback_id:
            row[4].value = "已回复"
            row[5].value = reply
            row[6].value = datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )
            break

    workbook.save(MANAGEMENT_PATH)


agent_component = components.declare_component(
    "k_company_engineering_agent",
    path=str(FRONTEND_DIR),
)


def main() -> None:
    st.set_page_config(
        page_title="K公司工程变更Agent",
        layout="wide",
        initial_sidebar_state="collapsed",
    )

    st.markdown(
        """
        <style>
        html, body, #root,
        [data-testid="stAppViewContainer"],
        [data-testid="stAppViewContainer"] > .main,
        section.main,
        section.main > div,
        [data-testid="stAppViewContainer"] > .main > div {
            margin: 0 !important;
            padding: 0 !important;
            width: 100% !important;
            max-width: none !important;
            min-height: 100vh;
            height: 100vh;
            overflow: hidden !important;
        }

        [data-testid="stVerticalBlock"],
        .stVerticalBlock,
        .stElementContainer,
        .stCustomComponentV1 {
            gap: 0 !important;
            margin: 0 !important;
            padding: 0 !important;
            width: 100% !important;
            max-width: none !important;
        }

        [data-testid="stHeader"],
        [data-testid="stToolbar"],
        [data-testid="stSidebar"],
        footer {
            display: none;
        }

        .stApp {
            background: #020617;
        }

        .block-container {
            max-width: none;
            padding: 0 !important;
        }

        iframe {
            display: block !important;
            width: 100% !important;
            max-width: none !important;
            height: 100vh !important;
            min-height: 100vh !important;
            margin: 0 !important;
            padding: 0 !important;
            border: 0 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    st.session_state.setdefault("role", "A")
    st.session_state.setdefault("response", "")
    st.session_state.setdefault("last_event", None)
    st.session_state.setdefault("conversation_history", [])
    st.session_state.setdefault("feedback_status", "")
    st.session_state.setdefault("response_token", -1)

    event = agent_component(
        role=st.session_state["role"],
        response=st.session_state["response"],
        response_token=st.session_state["response_token"],
        reference_files=load_reference_files(),
        versions=load_versions(),
        feedback=load_feedback(),
        feedback_status=st.session_state["feedback_status"],
        weather=load_weather(),
        default=None,
        key="agent_frontend",
    )

    if not event:
        return

    if event == st.session_state["last_event"]:
        return

    role = str(
        event.get(
            "role",
            st.session_state["role"],
        )
    ).upper()

    if role not in VALID_ROLES:
        return

    action = event.get("action", "ask")
    message = str(event.get("message", "")).strip()

    if action == "submit_feedback":
        append_feedback(role, message)
        st.session_state["feedback_status"] = "建议已提交"
        st.session_state["response"] = ""
        st.session_state["last_event"] = event
        st.rerun()

    if action == "reply_feedback":
        reply_feedback(
            str(event["feedback_id"]),
            message,
        )
        st.session_state["feedback_status"] = "回复已保存"
        st.session_state["response"] = ""
        st.session_state["last_event"] = event
        st.rerun()

    if action == "switch_role":
        st.session_state["role"] = role
        st.session_state["response"] = ""
        st.session_state["conversation_history"] = []
        st.session_state["last_event"] = event
        st.rerun()

    if not message:
        return

    history = st.session_state["conversation_history"]

    if action == "regenerate" and len(history) >= 2:
        history = history[:-2]

    try:
        interaction = handle_input(
            message,
            role,
            conversation_history=history,
        )
    except RuntimeError as error:
        st.session_state["response"] = str(error)
        st.session_state["response_token"] = int(
            event["request_token"]
        )
        st.session_state["last_event"] = event
        st.rerun()

    st.session_state["role"] = interaction["role"]
    st.session_state["response"] = interaction["answer"]
    st.session_state["response_token"] = int(
        event["request_token"]
    )

    history.append(
        {
            "role": "user",
            "content": message,
        }
    )

    history.append(
        {
            "role": "assistant",
            "content": interaction["answer"],
        }
    )

    st.session_state["conversation_history"] = history
    st.session_state["last_event"] = event

    st.rerun()


if __name__ == "__main__":
    main()
